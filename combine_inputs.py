import os
import openpyxl


def join_paths(path_1, path_2):
    return '\\'.join([path_1, path_2])


def copy_data(copy_from, copy_to):
    
    for row in copy_from.iter_rows(min_row = copy_from.min_row + 1, max_col = copy_from.max_column):
        copy_to.append([cell.value for cell in row])
    
    return copy_to


def main(file_directory, found_name):
    
    for file in os.listdir(file_directory):
        if found_name in file:
            current_template = openpyxl.load_workbook(join_paths(file_directory, file))
            current_template_sheet = current_template.active

    for file in os.listdir(file_directory):        
        if '.' not in file:
            particular_dir = join_paths(file_directory, file)
            for iner_file in os.listdir(particular_dir):
                try:
                    if found_name in iner_file:
                        current_file = openpyxl.load_workbook(join_paths(particular_dir, iner_file)).active
                        copy_data(current_file, current_template_sheet)
                        current_template.save(join_paths(file_directory, '2018 {} inputs.xlsx'.format(found_name)))
                except NameError:
                    print('File not found')
    return 


file_directory = input()
file_names_list = ['Preliminary', 'Actual', 'YoY']

for each in file_names_list:
    
    main(file_directory, each)

print('finish:')
