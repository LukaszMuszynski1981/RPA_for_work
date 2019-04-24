import openpyxl
import re
import time
import datetime as dt

start = time.time()

last_year_winspec = openpyxl.load_workbook(input())
this_year_winspec = openpyxl.load_workbook(input(), keep_vba=True)
gps = openpyxl.load_workbook(input())['2018 Fulfillment Audit']
mem_con_file = openpyxl.load_workbook(input() + r'\RAW_MEM_CON_FILE.xlsx')['DATA']


def clear_market_name(v):
    
    try:
        year_pattern = re.findall(r'\d{4}', v)[0].strip()
        v = ' '.join([i for i in v.split(' ') if i != year_pattern and i != ''])
    except IndexError:
        year_pattern = str(dt.datetime.today().year)
    
    return year_pattern, v


def extract_ctry_survey_code(v):
    
    v = clear_market_name(v)[1]
    ctry_code = re.findall(r'\s?[A-Z]{2}\s?', v)[0].strip()

    if re.findall(r'\s?[B]{1}\D{6}\s?', v):
        survey_code = 'BKG'
    elif re.findall(r'\s[A-Z]{3}\s?',v):
        survey_code = re.findall(r'\s[A-Z]{3}\s?',v)[0].strip()
    else:
        survey_code = 'XXX'

    return ctry_code, survey_code, v


def get_dates(market_view_name):
    
    for cel in gps['A:A']:
        if cel.value:
            headers = cel.row
            break

    for cel in gps[headers]:
        if cel.value == 'Item Abbreviated Name':
            survey_name_index = cel.col_idx - 1
        if cel.value == 'Product Region':
            region = cel.col_idx - 1
        if cel.value == 'Data Effective Date':
            effective_date_index = cel.col_idx - 1
        if cel.value == 'Planned Client Delivery Date (Client Delivery Date on extract)':
            publication_date_index = cel.col_idx - 1
        if cel.value == 'Production Team':
            production_team_index = cel.col_idx - 1

    for row in gps:
        try:
            if row[region].value == 'Latin America' and row[production_team_index].value == 'GOSS Warsaw LA Production' \
                    and clear_market_name(row[survey_name_index].value)[1] == market_view_name:
                final_market_view_name = row[survey_name_index].value
                effective_date = row[effective_date_index].value.date()
                publication_date = row[publication_date_index].value.date()
                break

            else:
                publication_date = dt.datetime.today()
                effective_date = publication_date
                final_market_view_name = 'Invalid market name - check Glob. Pub. Sched.'

        except TypeError:
            publication_date = dt.datetime.today()
            effective_date = publication_date
            final_market_view_name = 'Invalid market name - check Glob. Pub. Sched.'

    return effective_date, publication_date, final_market_view_name

'''fill the properties tab'''
lyws_properties = last_year_winspec['Properties']
tyws_properties = this_year_winspec['Properties']
lyws_properties_dict = {}        

for column in lyws_properties['A:A']:

    if column.value == 'Define Regression Bands':
        for i in range(0,3):
            lyws_properties_dict[column.offset(i,2).value] = column.offset(i, 3).value

    if column.value == 'Market View Publication Name':

        extract_ctry_survey_code_data = extract_ctry_survey_code(column.offset(0,2).value)
        ctry_code = extract_ctry_survey_code_data[0]
        survey_code = extract_ctry_survey_code_data[1]
        market_view_name = extract_ctry_survey_code_data[2]

        get_dates_data = get_dates(market_view_name)
        effective_date = get_dates_data[0]
        publication_date = get_dates_data[1]
        final_market_view_name = get_dates_data[2]

        lyws_properties_dict[column.value] = final_market_view_name

    if column.value == 'Translated Market View Publication Name':
        lyws_properties_dict[column.value] = final_market_view_name

    if column.value is None and column.value not in ['Define Regression Bands', 'Market View Publication Name',
                                                     'Translated Market View Publication Name']:
        lyws_properties_dict[column.value] = column.offset(0, 2).value


for column in tyws_properties['A:A']:

    if column.value == 'Market View Effective Date':
        column.offset(0, 2).value = effective_date.strftime('%A,%B %d, %Y')

    if column.value == 'Target Publish Date':
        column.offset(0, 2).value = publication_date.strftime('%A,%B %d, %Y')

    if column.value is None and column.value not in ['Define Regression Bands', 'Market View Effective Date',
                                                     'Target Publish Date']:
        if column.value in lyws_properties_dict.keys():
            column.offset(0,2).value = lyws_properties_dict[column.value]

    if column.value == 'Define Regression Bands':
        for i in range(0,3):
            if column.offset(i,2).value in lyws_properties_dict.keys():
                column.offset(i, 3).value = lyws_properties_dict[column.offset(i,2).value]
                    
'''fill the incumbent tab and fill the members from member configuration file'''  

lyws_incumbents = last_year_winspec['Incumbent Data']
tyws_incumbents = this_year_winspec['Incumbent Data']
ly_member_list = []
mat_member_list = []

for row in lyws_incumbents.rows:
    for cel in row:
        if cel.value is not None:
            if cel.value == 'Selected Member':
                select_index = cel.col_idx - 1
            if cel.value == 'WIN Display Label':
                member_name_index = cel.col_idx - 1
            try:
                if 'Detailed Report Members' in cel.value:
                    win_mat_index = cel.col_idx - 1
                if select_index and member_name_index and win_mat_index:
                    break
            except (NameError, TypeError):
                pass
                
for row in lyws_incumbents:
    if row[select_index].value in ['g', 'G', 'x', 'X'] and row[member_name_index].value not in ['Function', 'Job Family']:
        if row[member_name_index].value == 'Sales/Blue Collars/Others':
            ly_member_list.append('Sales/Blue Collars/Others/President')
        else:
            ly_member_list.append(row[member_name_index].value)
        
for row in mem_con_file.rows:
    for cel in row:
        if cel.value == 'WIN_MEMBER':
            mat_member_index = cel.col_idx - 1
        if cel.value == ctry_code:
            ctry_code_index = cel.col_idx - 1
        if mat_member_index and ctry_code_index:
            break

for row in mem_con_file:
    if row[ctry_code_index].value == 'X':
        mat_member_list.append(row[mat_member_index].value)
        
for row in tyws_incumbents:
    if row[member_name_index].value in ly_member_list:
        if row[select_index].value != 'G':
            row[select_index].value = 'X'
    if row[member_name_index].value in mat_member_list:
        row[win_mat_index].value = mat_member_list.index(row[member_name_index].value) + 1
     
        
'''save the final file'''
path_to_save = input()
this_year_winspec.save(path_to_save + r'\WIN_Specification_File_2019_test.xlsm')

stop = time.time()
print(stop-start)
